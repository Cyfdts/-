#2022/3/9
#@20计科kuku
import requests,time,smtplib,sys
from openpyxl import load_workbook
from email.mime.text import MIMEText
from lxml import etree
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from email.utils import formataddr
from selenium.webdriver.chrome.options import Options

def login(zh,mm):
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disabale-gpu")
    web = Chrome(options=chrome_options)
    # web = Chrome()
    web.get('https://tuan.12355.net/')

    time.sleep(1)

    web.find_element_by_xpath('//*[@id="userName"]').send_keys(f'{zh}') #账号密码

    web.find_element_by_xpath('//*[@id="password"]').send_keys(f'{mm}')
    time.sleep(1)

    web.find_element_by_xpath('//*[@id="login"]').click()  # 点击登录
    time.sleep(2)
    print("登录成功")
    web.find_element_by_xpath('//*[@id="nav"]/div[9]/div[1]/div[1]').click()
    time.sleep(4)
    windows = web.window_handles
    web.switch_to.window(windows[-1])
    web.find_element_by_xpath('/html/body/div/div/div[2]/div[1]/div[1]').click()
    time.sleep(1)
    web.find_element_by_xpath('/html/body/div/div/div[2]/div[1]/div/ul/div[2]/a/li').click()
    time.sleep(2)
    web.find_element_by_xpath('//*[@id="app"]/div/div[2]/section/div/div[5]/div[3]/table/tbody/tr/td[5]/div/div').click()
    time.sleep(1)
    tag1 = web.find_element_by_xpath('//*[@id="app"]/div/div[2]/section/div/div[5]/div/span[3]/div/input')
    maxpage = tag1.get_attribute("max")
    time.sleep(2)
    print("获取已完成青年大学习的团员名单中...")
    name_list = []
    for i in range(int(maxpage)):
        if(i>=1):
            web.find_element_by_xpath('//*[@id="app"]/div/div[2]/section/div/div[5]/div/button[2]').click()
            time.sleep(1)
        name_tag_list = web.find_elements_by_xpath('//*[@class="el-table__row"]//td[1]')
        situation_tag_list = web.find_elements_by_xpath('//*[@class="el-table__row"]//td[4]')
        for i in range(0,len(name_tag_list)):
            if(situation_tag_list[i].text=='已学'):
                name_list.append(name_tag_list[i].text)
    return name_list

def email_list_get(name_list):
    notdonename_list = []
    workfile = load_workbook('./附件1.xlsx')
    sheet = workfile.active
    print("从附件1.xlsx中对比出未完成青年大学习的团员名单")
    for i in range(1,len(sheet['A'])+1):
        if(sheet[f'A{i}'].value not in name_list):
            list1 = []
            list1.append(sheet[f'A{i}'].value),
            list1.append(sheet[f'B{i}'].value)
            notdonename_list.append(list1)
    return notdonename_list

def excel_msg_get():
    list2 = []
    workfile = load_workbook('./附件1.xlsx')
    sheet = workfile.active
    list2.append(sheet['D1'].value)
    list2.append(sheet['D2'].value)
    list2.append(sheet['D3'].value)
    list2.append(sheet['D4'].value)
    return list2

def email_send(name,email,from_addr,password):
    msg = MIMEText('青年大学习！不用回复！赶紧做！！')  # 构造邮件，内容为青年大学习
    msg["Subject"] = "青年大学习！！又是你！！！"  # 设置邮件主题
    msg["From"] = '辛苦勤奋艰苦奋斗的团支书'  # 寄件者
    msg["To"] = name  # 收件者
    smtp_server = 'smtp.qq.com'  # smtp服务器地址
    # to_addr = email  # 收件人地址
    try:
        # smtp协议的默认端口是25，QQ邮箱smtp服务器端口是465,第一个参数是smtp服务器地址，第二个参数是端口，第三个参数是超时设置,这里必须使用ssl证书，要不链接不上服务器
        server = smtplib.SMTP_SSL(smtp_server, 465, timeout=2)
        server.login(from_addr, password)  # 登录邮箱
        # 发送邮件，第一个参数是发送方地址，第二个参数是接收方列表，列表中可以有多个接收方地址，表示发送给多个邮箱，msg.as_string()将MIMEText对象转化成文本
        server.sendmail(from_addr, email, msg.as_string())
        server.quit()
        print(f"{name}的提醒通知已发送至{email}")
    except Exception as e:
        print('发送邮件失败: ', e)

if __name__ == '__main__':
    name_list1 = []
    try:
        zh = excel_msg_get()[0]
        mm = excel_msg_get()[1]
        from_addr = excel_msg_get()[2]
        password = excel_msg_get()[3]
    except Exception as t:
        print("excel格式有误,请检查！")
    try:
        name_list1 = login(zh,mm)
    except Exception as t:
        print("网络延迟较大,请重新运行脚本")
    if(len(name_list1)>=1):
        anounce_list = email_list_get(name_list1)
        print(f"还未完成当期青年大学习的同学有:{anounce_list}\n")
        sit = input("是否需要邮件通知他们？(请输入是或否)")
        if (sit == "是"):
            print("邮件开始发送")
            for i in anounce_list:
                if(i[1]==None):continue
                email_send(i[0], i[1], from_addr, password)
            print("所有邮件发送完毕")
    else:
        sys.exit()

